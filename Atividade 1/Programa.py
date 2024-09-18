'''
Atividade 1 Análise de Algorítmos - 17/09
Nome: Pedro Kourly
Turma: 6ºP Ciência da Computação
Professor: André Chaves

Criar dois programa para multiplicação de matrizes, o programa tradicional (utilizando duas estrutura de repetições) e o programa otimizado (utilizando apenas uma estrutura de repetição),  e executa-los para a mesma entrada n (n é a dimensões das matrizes. exemplo  matriz3x3 a dimensão é n= 3). 

Você deve pegar o tempo de execução de ambos os algoritmos para cada entrada n, fazer até n = 10 (ou seja matriz 10x10), e gerar uma tabela e um gráfico comparando as execução em relação de tempo x entrada.

---------------------------------------------------

O código irá salvar um relatório no arquivo "relatório.txt", contendo informações de tempo e passo a passo das iterações.
Além disso, a cada rodada, irá gerar um PNG com o gráfico contendo as informações de performances
'''
# Importa as bibliotecas necessárias
import numpy as np
import time
import matplotlib.pyplot as plt
from tqdm import tqdm
import openpyxl

def rand_matrix(n, min_val = 1, max_val = 100):
    # Essa função retorna uma matriz vazia de ordem 'n', com limites definidos como padrão 0 e 20.
    
    matrix = np.random.randint(min_val, max_val, size=(n, n))
    return matrix


# Problema o(n³)
def traditional_multiplication(n):
    # Essa função retorna o produto entre 2 matrizes aleatórias de mesma ordem n

    with open(f"relatório.txt", "a") as file:
        file.write("""Multiplicação tradicional [O(n³)]\n==================================""")

    # Variável que guardará o tempo
    execution_times = []

    for ordem in tqdm(range(1, n + 1), desc = 'Matriz tradicional'):
        # Criação das matrizes
        matrizA = rand_matrix(ordem)
        matrizB = rand_matrix(ordem)


        # Matriz resultado
        matrizRes = np.zeros((ordem, ordem))

        # O produto
        startTime = time.perf_counter()
        for i in range(ordem):
            for j in range(ordem):
                for k in range(ordem):
                    matrizRes[i][j] += matrizA[i][k] * matrizB[k][j]

        finalTime = time.perf_counter()

        execution_times.append(finalTime - startTime)

        # Gerá um relatório para cada iteração
        report_content = f"""
Report for Matrix Size {ordem}x{ordem}
=====================================

Matrix A:
{matrizA}

Matrix B:
{matrizB}

Multiplication Result:
{matrizRes}

Execution Time: {finalTime - startTime:.5f} seconds\n
"""

        with open(f"relatório.txt", "a") as file:
            file.write(report_content)
        
    return execution_times



def optimized_multiplication(n):
    with open(f"relatório.txt", "a") as file:
        file.write("""Multiplicação otimizada [O(n^2.81)]\n==================================""")

    execution_times = []

    for ordem in tqdm(range(1, n + 1), desc='Matriz otimizada'):
        # Criação das matrizes
        matrizA = rand_matrix(ordem)
        matrizB = rand_matrix(ordem)

        # Matriz resultado
        matrizRes = np.zeros((ordem, ordem))

        startTime = time.perf_counter()
        matrizRes = np.dot(matrizA, matrizB)  # Utiliza a multiplicação de matrizes do NumPy
        finalTime = time.perf_counter()

        execution_times.append(finalTime - startTime)

        report_content = f"""
Report for Matrix Size {ordem}x{ordem}
=====================================

Matrix A:
{matrizA}

Matrix B:
{matrizB}

Multiplication Result:
{matrizRes}

Execution Time: {finalTime - startTime:.5f} seconds\n
"""

        with open(f"relatório.txt", "a") as file:
            file.write(report_content)

    return execution_times


def main():
    report_content = """
Relatório de atividade\n

Atividade 1 Análise de Algorítmos
Nome: Pedro Kourly
Turma: 6ºP Ciência da Computação
Professor: André Chaves
===================================
          """
    
    with open(f"relatório.txt", "w") as file:
        file.write(report_content)

    # Inicializa as variáveis do algorítmo
    n = 10

    traditional_times = traditional_multiplication(n)
    optimized_times = optimized_multiplication(n)

    print(traditional_times)
    print(optimized_times)
    # Gerar gráfico
    plt.plot(range(1, n + 1), traditional_times, label='Multiplicação Tradicional')
    plt.plot(range(1, n + 1), optimized_times, label='Multiplicação Otimizada')
    plt.xlabel('Tamanho da Matriz (n)')
    plt.ylabel('Tempo de Execução (s)')
    plt.title('Comparação de Performance')
    plt.legend()
    plt.savefig("Atividade_1.png")
    plt.show()
    
    # Escrever tempos de execução em um arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tempos de Execução"
    ws['A1'] = "n"
    ws['B1'] = "Tempo Tradicional (s)"
    ws['C1'] = "Tempo Otimizado (s)"
    for i in range(1, n + 1):
        ws[f'A{i+1}'] = i
        ws[f'B{i+1}'].value = float(format(traditional_times[i-1], '.15f'))
        ws[f'B{i+1}'].number_format = '_(* #,##0.0000000000000000_);_(* (#,##0.0000000000000000);_(* "-"??_);_(@_)'
        ws[f'C{i+1}'].value = float(format(optimized_times[i-1], '.15f'))
        ws[f'C{i+1}'].number_format = '_(* #,##0.0000000000000000_);_(* (#,##0.0000000000000000);_(* "-"??_);_(@_)'
    wb.save("tempos_execucao.xlsx")

if __name__ == '__main__':
    main()